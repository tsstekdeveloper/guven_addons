import base64
import io
import logging
from datetime import datetime

from odoo import api, fields, models
from odoo.exceptions import UserError

_logger = logging.getLogger(__name__)

# ================================================================
# Date formats
# ================================================================

DATE_FORMATS = [
    '%Y%m%d',              # QNB primary: 20240115
    '%Y-%m-%d %H:%M:%S',
    '%d/%m/%Y',
    '%d-%m-%Y',
    '%d.%m.%Y',
    '%Y-%m-%d',
]

# ================================================================
# Column positions
# ================================================================

# Common columns (0-21) — shared by E-Fatura and E-Arsiv sheets
_COMMON = {
    'uuid': 0,
    'profile_id': 1,
    'invoice_id': 2,
    'issue_date': 3,
    'sender': 5,
    'sender_name': 6,
    'receiver': 7,
    'receiver_name': 8,
    'payable_amount': 9,
    'currency_code': 10,
    'invoice_type_code': 11,
    'tax_exclusive_amount': 16,
    'tax_inclusive_amount': 18,
    'allowance_total_amount': 20,
}

# E-Fatura (81 cols): KDV (rate, matrah_col, tutar_col)
_EFATURA_KDV = [
    (0, 28, 30), (1, 32, 34), (8, 36, 38), (9, 40, 42),
    (10, 44, 46), (18, 48, 50), (20, 52, 54),
]
_EFATURA_EXTRA = {
    'exchange_rate': 79,
    'kdv_tevkifat_matrah': 60,
    'kdv_tevkifat_tutar': 65,
    'otv_tevkifat_matrah': 62,
    'otv_tevkifat_tutar': 67,
    'oiv_matrah': 69,
    'oiv_tutar': 71,
}

# E-Arsiv (74 cols): KDV (rate, matrah_col, tutar_col) — no %9
_EARSIV_KDV = [
    (0, 28, 30), (1, 32, 34), (8, 36, 38),
    (10, 40, 42), (18, 44, 46), (20, 48, 50),
]
_EARSIV_EXTRA = {
    'exchange_rate': 73,
    'kdv_tevkifat_matrah': 56,
    'kdv_tevkifat_tutar': 60,
    'otv_tevkifat_matrah': 58,
    'otv_tevkifat_tutar': 62,
    'oiv_matrah': 64,
    'oiv_tutar': 66,
}

# Valid profile_id / invoice_type_code values (must match guven_fatura.py)
_VALID_PROFILES = {
    'TEMELFATURA', 'TICARIFATURA', 'IHRACATFATURA', 'YOLCUBERABERFATURA',
    'EARSIVFATURA', 'ILAC_TIBBICIHAZ', 'ENERJI', 'YATIRIMTESVIK',
    'KONAKLAMAVERGISI', 'HKS', 'KAMU', 'KAMUFATURA', 'SGK', 'IDIS', 'IPTAL',
}
_VALID_TYPES = {
    'SATIS', 'IADE', 'TEVKIFAT', 'TEVKIFATIADE', 'ISTISNA', 'OZELMATRAH',
    'IHRACKAYITLI', 'IHRACAT', 'SARJ', 'SARJANLIK', 'YTBSATIS', 'YTBISTISNA',
    'YTBIADE', 'KOMISYONCU', 'SGK',
}


# ================================================================
# Helpers
# ================================================================

def _parse_excel_date(value):
    """Parse various date formats from QNB Excel."""
    if not value:
        return False
    if isinstance(value, datetime):
        return value.date()
    if hasattr(value, 'date'):
        d = value.date
        return d() if callable(d) else d
    val_str = str(value).strip()
    if not val_str:
        return False
    for fmt in DATE_FORMATS:
        try:
            return datetime.strptime(val_str, fmt).date()
        except (ValueError, TypeError):
            continue
    _logger.warning("QNB date parse failed: %s", val_str)
    return False


def _safe_float(value):
    """Convert Excel cell value to float safely.

    Handles both formats:
      English: 1234.56   (dot = decimal)
      Turkish: 1.234,56  (dot = thousands, comma = decimal)
    Detection: if comma present → Turkish, otherwise English.
    """
    if value is None:
        return 0.0
    if isinstance(value, (int, float)):
        return float(value)
    val_str = str(value).strip()
    if not val_str:
        return 0.0
    if ',' in val_str:
        # Turkish format: 1.234,56 → 1234.56
        val_str = val_str.replace('.', '').replace(',', '.')
    try:
        return float(val_str)
    except (ValueError, TypeError):
        return 0.0


def _normalize_vkn(value):
    """Normalize VKN/TCKN for comparison and storage.

    Excel may store VKN as float (4510016851.0) or with TR prefix.
    """
    if value is None:
        return ''
    if isinstance(value, float):
        value = int(value)
    val = str(value).strip()
    if val.endswith('.0'):
        val = val[:-2]
    if val.upper().startswith('TR'):
        val = val[2:]
    return val.strip()


def _cell(row, col, shift=0):
    """Get cell value with optional column shift (e-arsiv col >= 23)."""
    actual = col - shift if col >= 23 else col
    if 0 <= actual < len(row):
        return row[actual]
    return None


def _is_numeric(value):
    """Check if a cell value is numeric."""
    if value is None:
        return False
    if isinstance(value, (int, float)):
        return True
    try:
        float(str(value).strip().replace('.', '').replace(',', '.'))
        return True
    except (ValueError, TypeError):
        return False


# ================================================================
# TransientModels
# ================================================================

class GuvenQnbImportFile(models.TransientModel):
    _name = 'guven.qnb.import.file'
    _description = 'QNB Excel Dosyasi'

    wizard_id = fields.Many2one(
        'guven.qnb.import.wizard', ondelete='cascade',
    )
    file_data = fields.Binary('Excel Dosyasi', required=True)
    file_name = fields.Char('Dosya Adi')
    efatura_count = fields.Integer('E-Fatura', readonly=True)
    earsiv_count = fields.Integer('E-Arsiv', readonly=True)
    status = fields.Char('Durum', readonly=True)


class GuvenQnbImportWizard(models.TransientModel):
    _name = 'guven.qnb.import.wizard'
    _description = 'QNB Excel Iceri Al'

    company_id = fields.Many2one(
        'res.company', string='Sirket', required=True,
        default=lambda self: self.env.company, readonly=True,
    )
    company_vat = fields.Char(related='company_id.vat', string='VKN')
    file_ids = fields.One2many(
        'guven.qnb.import.file', 'wizard_id', string='Excel Dosyalari',
    )
    auto_logo_sync = fields.Boolean('Logo Eslestirmesi Yap', default=True)
    state = fields.Selection(
        [('draft', 'Taslak'), ('preview', 'On Izleme'), ('done', 'Tamamlandi')],
        default='draft',
    )
    import_summary = fields.Html(readonly=True)
    preview_efatura_total = fields.Integer(readonly=True)
    preview_earsiv_total = fields.Integer(readonly=True)
    preview_new = fields.Integer(readonly=True)
    preview_update = fields.Integer(readonly=True)
    preview_unchanged = fields.Integer(readonly=True)
    preview_skip = fields.Integer(readonly=True)

    # ================================================================
    # Helpers
    # ================================================================

    def _read_workbook(self, file_record):
        """Read an openpyxl workbook from a file record."""
        try:
            import openpyxl
        except ImportError:
            raise UserError(
                "openpyxl kutuphanesi yuklu degil. "
                "Lutfen 'pip install openpyxl' komutuyla yukleyin."
            )
        raw = base64.b64decode(file_record.file_data)
        return openpyxl.load_workbook(io.BytesIO(raw), data_only=True)

    def _validate_vkn(self, ws, file_name, sheet_label):
        """Validate that first 10 data rows contain matching company VKN."""
        company_vat = _normalize_vkn(self.company_id.vat)
        if not company_vat:
            raise UserError(
                f"'{self.company_id.name}' sirketinde VKN tanimlanmamis. "
                "Lutfen sirket ayarlarindan VKN girin."
            )
        for row in ws.iter_rows(min_row=2, max_row=11, values_only=True):
            if not row or len(row) < 9:
                continue
            sender_vkn = _normalize_vkn(row[5])
            receiver_vkn = _normalize_vkn(row[7])
            if sender_vkn == company_vat or receiver_vkn == company_vat:
                return True
        raise UserError(
            f"'{file_name}' dosyasinin {sheet_label} sayfasinda "
            f"sirket VKN'si ({company_vat}) bulunamadi.\n"
            "Lutfen dogru sirkette oldugunuzdan emin olun."
        )

    def _detect_earsiv_shift(self, ws):
        """Detect column shift in e-arsiv sheet by checking col 28."""
        for row in ws.iter_rows(min_row=2, max_row=5, values_only=True):
            if not row or len(row) <= 28:
                continue
            if not _is_numeric(row[28]):
                _logger.info("QNB e-arsiv: column shift detected (col 28 non-numeric)")
                return 1
            return 0
        return 0

    def _determine_direction(self, sender_vkn, receiver_vkn):
        """Determine invoice direction based on VKN match with company."""
        company_vat = _normalize_vkn(self.company_id.vat)
        if _normalize_vkn(sender_vkn) == company_vat:
            return 'OUT'
        return 'IN'

    def _parse_sheet(self, ws, kdv_cols, extra_cols, kaynak, shift=0):
        """Parse all data rows from a QNB Excel sheet.

        Returns list of dicts with invoice data + nested 'taxes' list.
        """
        rows = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            if not row or not any(row):
                continue

            # Common fields
            uuid = str(row[_COMMON['uuid']] or '').strip()
            invoice_id = str(row[_COMMON['invoice_id']] or '').strip()
            if not uuid or not invoice_id:
                continue

            currency = str(row[_COMMON['currency_code']] or 'TRY').strip()
            if not currency:
                currency = 'TRY'

            data = {
                'uuid': uuid,
                'invoice_id': invoice_id,
                'profile_id': str(row[_COMMON['profile_id']] or '').strip(),
                'issue_date': _parse_excel_date(row[_COMMON['issue_date']]),
                'sender': _normalize_vkn(row[_COMMON['sender']]),
                'sender_name': str(row[_COMMON['sender_name']] or '').strip(),
                'receiver': _normalize_vkn(row[_COMMON['receiver']]),
                'receiver_name': str(row[_COMMON['receiver_name']] or '').strip(),
                'payable_amount': _safe_float(row[_COMMON['payable_amount']]),
                'currency_code': currency,
                'invoice_type_code': str(row[_COMMON['invoice_type_code']] or '').strip(),
                'tax_exclusive_amount': _safe_float(row[_COMMON['tax_exclusive_amount']]),
                'tax_inclusive_amount': _safe_float(row[_COMMON['tax_inclusive_amount']]),
                'allowance_total_amount': _safe_float(row[_COMMON['allowance_total_amount']]),
                'kaynak': kaynak,
            }

            # Exchange rate (columns >= 23 may be shifted in e-arsiv)
            data['exchange_rate'] = (
                _safe_float(_cell(row, extra_cols['exchange_rate'], shift)) or 1.0
            )

            # --- Tax entries ---
            taxes = []

            # KDV rates
            for rate, matrah_col, tutar_col in kdv_cols:
                matrah = _safe_float(_cell(row, matrah_col, shift))
                tutar = _safe_float(_cell(row, tutar_col, shift))
                if matrah or tutar:
                    taxes.append({
                        'tax_type': 'kdv',
                        'percent': float(rate),
                        'taxable_amount': matrah,
                        'tax_amount': tutar,
                    })

            # KDV Tevkifat
            tev_matrah = _safe_float(_cell(row, extra_cols['kdv_tevkifat_matrah'], shift))
            tev_tutar = _safe_float(_cell(row, extra_cols['kdv_tevkifat_tutar'], shift))
            if tev_matrah or tev_tutar:
                taxes.append({
                    'tax_type': 'withholding',
                    'percent': 0.0,
                    'taxable_amount': tev_matrah,
                    'tax_amount': tev_tutar,
                })

            # OTV Tevkifat — not supported, raise error if data found
            otv_matrah = _safe_float(_cell(row, extra_cols['otv_tevkifat_matrah'], shift))
            otv_tutar = _safe_float(_cell(row, extra_cols['otv_tevkifat_tutar'], shift))
            if otv_matrah or otv_tutar:
                raise UserError(
                    f"Fatura {invoice_id}: OTV Tevkifat verisi tespit edildi "
                    f"(matrah={otv_matrah}, tutar={otv_tutar}). "
                    "OTV Tevkifat henuz desteklenmiyor."
                )

            # OIV (Ozel Iletisim Vergisi)
            oiv_matrah = _safe_float(_cell(row, extra_cols['oiv_matrah'], shift))
            oiv_tutar = _safe_float(_cell(row, extra_cols['oiv_tutar'], shift))
            if oiv_matrah or oiv_tutar:
                taxes.append({
                    'tax_type': 'oiv',
                    'percent': 0.0,
                    'taxable_amount': oiv_matrah,
                    'tax_amount': oiv_tutar,
                })

            data['taxes'] = taxes
            rows.append(data)

        return rows

    def _build_vals(self, row_data, company):
        """Build guven.fatura create/write values from parsed row data."""
        currency = (row_data['currency_code'] or 'TRY').upper()
        is_try = currency == 'TRY'
        exchange_rate = row_data.get('exchange_rate', 1.0) or 1.0

        profile = row_data['profile_id']
        inv_type = row_data['invoice_type_code']

        vals = {
            'uuid': row_data['uuid'],
            'invoice_id': row_data['invoice_id'],
            'profile_id': profile if profile in _VALID_PROFILES else False,
            'invoice_type_code': inv_type if inv_type in _VALID_TYPES else False,
            'issue_date': row_data['issue_date'],
            'sender': row_data['sender'],
            'sender_name': row_data['sender_name'],
            'receiver': row_data['receiver'],
            'receiver_name': row_data['receiver_name'],
            'payable_amount': row_data['payable_amount'],
            'currency_code': currency,
            'exchange_rate': exchange_rate,
            'tax_exclusive_amount': row_data['tax_exclusive_amount'],
            'tax_inclusive_amount': row_data['tax_inclusive_amount'],
            'allowance_total_amount': row_data['allowance_total_amount'],
            'kaynak': row_data['kaynak'],
            'direction': self._determine_direction(
                row_data['sender'], row_data['receiver'],
            ),
            'details_received': True,
            'is_cancellation': False,
            'company_id': company.id,
        }

        # TRY amounts
        if is_try:
            vals['tax_exclusive_amount_try'] = row_data['tax_exclusive_amount']
            vals['tax_inclusive_amount_try'] = row_data['tax_inclusive_amount']
            vals['allowance_total_amount_try'] = row_data['allowance_total_amount']
            vals['payable_amount_try'] = row_data['payable_amount']
        else:
            vals['tax_exclusive_amount_try'] = row_data['tax_exclusive_amount'] * exchange_rate
            vals['tax_inclusive_amount_try'] = row_data['tax_inclusive_amount'] * exchange_rate
            vals['allowance_total_amount_try'] = row_data['allowance_total_amount'] * exchange_rate
            vals['payable_amount_try'] = row_data['payable_amount'] * exchange_rate

        return vals

    def _sync_tax_records(self, fatura, row_data):
        """Create/replace guven.fatura.tax records for a fatura."""
        Tax = self.env['guven.fatura.tax']
        currency = (row_data['currency_code'] or 'TRY').upper()
        is_try = currency == 'TRY'
        exchange_rate = row_data.get('exchange_rate', 1.0) or 1.0

        # Remove existing tax records, then recreate
        fatura.tax_ids.unlink()

        for td in row_data.get('taxes', []):
            taxable = td['taxable_amount']
            amount = td['tax_amount']
            vals = {
                'fatura_id': fatura.id,
                'tax_type': td['tax_type'],
                'percent': td['percent'],
                'taxable_amount': taxable,
                'tax_amount': amount,
                'currency_code': currency,
            }
            if is_try:
                vals['taxable_amount_try'] = taxable
                vals['tax_amount_try'] = amount
            else:
                vals['taxable_amount_try'] = taxable * exchange_rate
                vals['tax_amount_try'] = amount * exchange_rate
            Tax.create(vals)

    # ================================================================
    # Change detection
    # ================================================================

    # Fields to compare for change detection
    _COMPARE_FIELDS = (
        'invoice_id', 'profile_id', 'invoice_type_code', 'issue_date',
        'sender', 'sender_name', 'receiver', 'receiver_name',
        'payable_amount', 'currency_code', 'exchange_rate',
        'tax_exclusive_amount', 'tax_inclusive_amount',
        'allowance_total_amount', 'direction',
        'tax_exclusive_amount_try', 'tax_inclusive_amount_try',
        'allowance_total_amount_try', 'payable_amount_try',
    )
    _FLOAT_TOLERANCE = 0.005

    def _has_changes(self, existing, vals):
        """Check if vals differ from existing record."""
        for field in self._COMPARE_FIELDS:
            if field not in vals:
                continue
            new_val = vals[field]
            old_val = getattr(existing, field, None)
            if isinstance(new_val, float) or isinstance(old_val, float):
                old_f = float(old_val or 0)
                new_f = float(new_val or 0)
                if abs(old_f - new_f) > self._FLOAT_TOLERANCE:
                    return True
            else:
                if (new_val or False) != (old_val or False):
                    return True
        return False

    def _taxes_changed(self, existing, row_data):
        """Check if tax records differ from parsed data."""
        existing_taxes = existing.tax_ids
        new_taxes = row_data.get('taxes', [])
        if len(existing_taxes) != len(new_taxes):
            return True
        # Sort both by (tax_type, percent) for comparison
        old_sorted = sorted(existing_taxes, key=lambda t: (t.tax_type or '', t.percent or 0))
        new_sorted = sorted(new_taxes, key=lambda t: (t['tax_type'] or '', t['percent'] or 0))
        for old_t, new_t in zip(old_sorted, new_sorted):
            if old_t.tax_type != new_t['tax_type']:
                return True
            if abs((old_t.percent or 0) - (new_t['percent'] or 0)) > self._FLOAT_TOLERANCE:
                return True
            if abs((old_t.taxable_amount or 0) - (new_t['taxable_amount'] or 0)) > self._FLOAT_TOLERANCE:
                return True
            if abs((old_t.tax_amount or 0) - (new_t['tax_amount'] or 0)) > self._FLOAT_TOLERANCE:
                return True
        return False

    # ================================================================
    # Actions
    # ================================================================

    def action_preview(self):
        """Validate files and compute preview counts (draft -> preview)."""
        self.ensure_one()
        if not self.file_ids:
            raise UserError("Lutfen en az bir Excel dosyasi ekleyin.")

        Fatura = self.env['guven.fatura']
        company = self.company_id
        efatura_total = earsiv_total = new = update = unchanged = skip = 0

        for frec in self.file_ids:
            wb = self._read_workbook(frec)
            fname = frec.file_name or 'dosya'
            file_new = file_update = file_unchanged = file_skip = 0
            file_efatura = file_earsiv = 0

            # Sheet 0: E-Fatura (81 cols)
            if len(wb.worksheets) >= 1:
                ws = wb.worksheets[0]
                self._validate_vkn(ws, fname, 'E-Fatura')
                rows = self._parse_sheet(
                    ws, _EFATURA_KDV, _EFATURA_EXTRA, 'e-fatura-qnbexcel',
                )
                file_efatura = len(rows)
                for rd in rows:
                    existing = Fatura.search([
                        ('uuid', '=', rd['uuid']),
                        ('kaynak', '=', 'e-fatura-qnbexcel'),
                        ('company_id', '=', company.id),
                    ], limit=1)
                    if existing:
                        if existing.is_locked:
                            file_skip += 1
                        else:
                            vals = self._build_vals(rd, company)
                            if self._has_changes(existing, vals) or self._taxes_changed(existing, rd):
                                file_update += 1
                            else:
                                file_unchanged += 1
                    else:
                        file_new += 1

            # Sheet 1: E-Arsiv (74 cols)
            if len(wb.worksheets) >= 2:
                ws = wb.worksheets[1]
                self._validate_vkn(ws, fname, 'E-Arsiv')
                shift = self._detect_earsiv_shift(ws)
                rows = self._parse_sheet(
                    ws, _EARSIV_KDV, _EARSIV_EXTRA, 'e-arsiv-qnbexcel', shift,
                )
                file_earsiv = len(rows)
                for rd in rows:
                    existing = Fatura.search([
                        ('uuid', '=', rd['uuid']),
                        ('kaynak', '=', 'e-arsiv-qnbexcel'),
                        ('company_id', '=', company.id),
                    ], limit=1)
                    if existing:
                        if existing.is_locked:
                            file_skip += 1
                        else:
                            vals = self._build_vals(rd, company)
                            if self._has_changes(existing, vals) or self._taxes_changed(existing, rd):
                                file_update += 1
                            else:
                                file_unchanged += 1
                    else:
                        file_new += 1

            wb.close()

            frec.write({
                'efatura_count': file_efatura,
                'earsiv_count': file_earsiv,
                'status': (
                    f"{file_new} yeni, {file_update} guncelleme, "
                    f"{file_unchanged} degisiklik yok, {file_skip} atlama"
                ),
            })
            efatura_total += file_efatura
            earsiv_total += file_earsiv
            new += file_new
            update += file_update
            unchanged += file_unchanged
            skip += file_skip

        self.write({
            'state': 'preview',
            'preview_efatura_total': efatura_total,
            'preview_earsiv_total': earsiv_total,
            'preview_new': new,
            'preview_update': update,
            'preview_unchanged': unchanged,
            'preview_skip': skip,
        })
        return {
            'type': 'ir.actions.act_window',
            'res_model': self._name,
            'res_id': self.id,
            'view_mode': 'form',
            'target': 'new',
        }

    def action_import(self):
        """Execute import (preview -> done)."""
        self.ensure_one()
        Fatura = self.env['guven.fatura']
        company = self.company_id
        created = updated = unchanged = skipped = 0
        errors = []

        for frec in self.file_ids:
            wb = self._read_workbook(frec)

            # Collect all rows from both sheets
            all_rows = []
            if len(wb.worksheets) >= 1:
                ws = wb.worksheets[0]
                all_rows.extend(
                    self._parse_sheet(ws, _EFATURA_KDV, _EFATURA_EXTRA, 'e-fatura-qnbexcel')
                )
            if len(wb.worksheets) >= 2:
                ws = wb.worksheets[1]
                shift = self._detect_earsiv_shift(ws)
                all_rows.extend(
                    self._parse_sheet(
                        ws, _EARSIV_KDV, _EARSIV_EXTRA, 'e-arsiv-qnbexcel', shift,
                    )
                )
            wb.close()

            for rd in all_rows:
                try:
                    existing = Fatura.search([
                        ('uuid', '=', rd['uuid']),
                        ('kaynak', '=', rd['kaynak']),
                        ('company_id', '=', company.id),
                    ], limit=1)

                    vals = self._build_vals(rd, company)

                    if existing:
                        if existing.is_locked:
                            skipped += 1
                            continue
                        if not self._has_changes(existing, vals) and not self._taxes_changed(existing, rd):
                            unchanged += 1
                            continue
                        # Don't overwrite uuid on update
                        vals.pop('uuid', None)
                        existing.write(vals)
                        self._sync_tax_records(existing, rd)
                        updated += 1
                    else:
                        fatura = Fatura.create(vals)
                        self._sync_tax_records(fatura, rd)
                        created += 1
                except Exception as e:
                    _logger.exception(
                        "QNB import error for invoice %s",
                        rd.get('invoice_id', '?'),
                    )
                    errors.append(f"{rd.get('invoice_id', '?')}: {e}")

        # Build summary HTML
        summary_parts = [
            '<div class="alert alert-success">',
            '<strong>Import tamamlandi!</strong><br/>',
            f'Yeni: <strong>{created}</strong> | ',
            f'Guncellenen: <strong>{updated}</strong> | ',
            f'Degisiklik yok: <strong>{unchanged}</strong> | ',
            f'Atlanan (kilitli): <strong>{skipped}</strong>',
            '</div>',
        ]
        if errors:
            summary_parts.append('<div class="alert alert-danger">')
            summary_parts.append(f'<strong>{len(errors)} hata:</strong><ul>')
            for err in errors[:20]:
                summary_parts.append(f'<li>{err}</li>')
            if len(errors) > 20:
                summary_parts.append(f'<li>...ve {len(errors) - 20} hata daha</li>')
            summary_parts.append('</ul></div>')

        # Optional Logo sync — compute date range from imported invoices
        logo_msg = ''
        if self.auto_logo_sync and created + updated > 0:
            try:
                qnb_invoices = self.env['guven.fatura'].search([
                    ('kaynak', 'in', ['e-fatura-qnbexcel', 'e-arsiv-qnbexcel']),
                    ('company_id', '=', company.id),
                    ('issue_date', '!=', False),
                ])
                if qnb_invoices:
                    dates = qnb_invoices.mapped('issue_date')
                    logo_wizard = self.env['guven.logo.sync.wizard'].create({
                        'company_ids': [(6, 0, [company.id])],
                        'date_from': min(dates),
                        'date_to': max(dates),
                    })
                    logo_wizard.action_sync()
                    match_total = logo_wizard.match_total or 0
                    match_single = logo_wizard.match_single or 0
                    logo_msg = (
                        f'<div class="alert alert-info">'
                        f'<strong>Logo Eslestirme:</strong> '
                        f'{match_total} fatura tarandi, {match_single} eslesme bulundu.'
                        f'</div>'
                    )
            except Exception as e:
                _logger.exception("Logo sync after QNB import failed")
                logo_msg = (
                    f'<div class="alert alert-warning">'
                    f'<strong>Logo eslestirme hatasi:</strong> {e}'
                    f'</div>'
                )

        self.write({
            'state': 'done',
            'import_summary': ''.join(summary_parts) + logo_msg,
        })
        return {
            'type': 'ir.actions.act_window',
            'res_model': self._name,
            'res_id': self.id,
            'view_mode': 'form',
            'target': 'new',
        }

    def action_back(self):
        """Go back to draft state (preview -> draft)."""
        self.ensure_one()
        self.write({
            'state': 'draft',
            'preview_efatura_total': 0,
            'preview_earsiv_total': 0,
            'preview_new': 0,
            'preview_update': 0,
            'preview_unchanged': 0,
            'preview_skip': 0,
        })
        return {
            'type': 'ir.actions.act_window',
            'res_model': self._name,
            'res_id': self.id,
            'view_mode': 'form',
            'target': 'new',
        }

    def action_close(self):
        return {
            'type': 'ir.actions.client',
            'tag': 'reload',
        }
