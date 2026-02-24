import base64
import io
import logging
from datetime import datetime

from odoo import api, fields, models
from odoo.exceptions import UserError

_logger = logging.getLogger(__name__)

# GİB Excel header (Row 2) — normalized for comparison
REQUIRED_HEADERS = [
    'sıra', 'ünvanı/adı soyadı', 'vergi kimlik/t.c. kimlik numarası',
    'fatura no', 'düzenleme tarihi', 'toplam tutar', 'ödenecek tutar',
    'vergiler toplamı', 'para birimi', 'tesisat numarası',
    'gönderim şekli', 'iptal itiraz durum', 'iptal itiraz tarihi',
]

DATE_FORMATS = [
    '%Y-%m-%d %H:%M:%S',
    '%d/%m/%Y',
    '%d-%m-%Y',
    '%d.%m.%Y',
    '%Y-%m-%d',
]


def _parse_excel_date(value):
    """Parse various date formats from GİB Excel."""
    if not value:
        return False
    if isinstance(value, datetime):
        return value.date()
    if hasattr(value, 'date'):
        return value.date() if callable(value.date) else value.date
    val_str = str(value).strip()
    if not val_str:
        return False
    for fmt in DATE_FORMATS:
        try:
            return datetime.strptime(val_str, fmt).date()
        except (ValueError, TypeError):
            continue
    _logger.warning("Could not parse date: %s", val_str)
    return False


def _safe_float(value):
    """Convert Excel cell value to float safely."""
    if value is None:
        return 0.0
    if isinstance(value, (int, float)):
        return float(value)
    val_str = str(value).strip()
    if not val_str:
        return 0.0
    # Handle Turkish decimal format (1.234,56 → 1234.56)
    val_str = val_str.replace('.', '').replace(',', '.')
    try:
        return float(val_str)
    except (ValueError, TypeError):
        return 0.0


class GuvenEarsivImportFile(models.TransientModel):
    _name = 'guven.earsiv.import.file'
    _description = 'E-Arşiv GİB Excel Dosyası'

    wizard_id = fields.Many2one(
        'guven.earsiv.import.wizard', ondelete='cascade',
    )
    file_data = fields.Binary('Excel Dosyası', required=True)
    file_name = fields.Char('Dosya Adı')
    row_count = fields.Integer('Satır Sayısı', readonly=True)
    status = fields.Char('Durum', readonly=True)


class GuvenEarsivImportWizard(models.TransientModel):
    _name = 'guven.earsiv.import.wizard'
    _description = 'E-Arşiv GİB Excel İçeri Al'

    company_id = fields.Many2one(
        'res.company', string='Şirket', required=True,
        default=lambda self: self.env.company, readonly=True,
    )
    company_vat = fields.Char(related='company_id.vat', string='VKN')
    file_ids = fields.One2many(
        'guven.earsiv.import.file', 'wizard_id', string='Excel Dosyaları',
    )
    auto_logo_sync = fields.Boolean('Logo Eşleştirmesi Yap', default=True)
    state = fields.Selection(
        [('draft', 'Taslak'), ('preview', 'Ön İzleme'), ('done', 'Tamamlandı')],
        default='draft',
    )
    import_summary = fields.Html(readonly=True)
    preview_total = fields.Integer(readonly=True)
    preview_new = fields.Integer(readonly=True)
    preview_update = fields.Integer(readonly=True)
    preview_skip = fields.Integer(readonly=True)

    # ----------------------------------------------------------------
    # Helpers
    # ----------------------------------------------------------------

    def _read_workbook(self, file_record):
        """Read an openpyxl workbook from a file record."""
        try:
            import openpyxl
        except ImportError:
            raise UserError(
                "openpyxl kütüphanesi yüklü değil. "
                "Lütfen 'pip install openpyxl' komutuyla yükleyin."
            )
        raw = base64.b64decode(file_record.file_data)
        wb = openpyxl.load_workbook(io.BytesIO(raw), read_only=True, data_only=True)
        return wb

    @staticmethod
    def _normalize_turkish(text):
        """Normalize Turkish text for case-insensitive comparison.

        Handles Turkish İ/I dotted/dotless issue:
        Python's str.lower() turns İ (U+0130) into i + combining dot (U+0307),
        which doesn't match plain 'i'. We strip the combining dot after lowering.
        """
        text = str(text or '').strip().lower()
        # Remove combining dot above (U+0307) left over from İ → i̇
        return text.replace('\u0307', '')

    def _validate_headers(self, ws, file_name):
        """Validate that row 2 contains the expected GİB headers."""
        header_row = []
        for cell in ws[2]:
            val = self._normalize_turkish(cell.value)
            header_row.append(val)
        # Check that all required headers are present (order-based)
        for idx, expected in enumerate(REQUIRED_HEADERS):
            if idx >= len(header_row):
                raise UserError(
                    f"'{file_name}' dosyasında beklenen kolon sayısı ({len(REQUIRED_HEADERS)}) "
                    f"bulunamadı. Kolon {idx + 1}: '{expected}' eksik."
                )
            actual = header_row[idx]
            if expected not in actual:
                raise UserError(
                    f"'{file_name}' dosyasında kolon {idx + 1} hatası.\n"
                    f"Beklenen: '{expected}'\nBulunan: '{actual}'"
                )

    def _parse_rows(self, ws):
        """Parse data rows (row 3+) from a GİB Excel worksheet. Returns list of dicts."""
        rows = []
        for row in ws.iter_rows(min_row=3, values_only=True):
            # Skip empty rows
            if not row or not any(row):
                continue
            # Column mapping (0-indexed): A=0 Sıra, B=1 Ünvan, C=2 VKN, ...
            invoice_id = str(row[3] or '').strip() if len(row) > 3 else ''
            if not invoice_id:
                continue
            sender_vkn = str(row[2] or '').strip() if len(row) > 2 else ''
            sender_name = str(row[1] or '').strip() if len(row) > 1 else ''
            issue_date = _parse_excel_date(row[4] if len(row) > 4 else None)
            toplam_tutar = _safe_float(row[5] if len(row) > 5 else 0)
            odenecek_tutar = _safe_float(row[6] if len(row) > 6 else 0)
            vergiler_toplami = _safe_float(row[7] if len(row) > 7 else 0)
            para_birimi = str(row[8] or 'TRY').strip() if len(row) > 8 else 'TRY'
            iptal_durum = str(row[10] or '').strip() if len(row) > 10 else ''

            is_iptal = 'İptal' in iptal_durum or 'iptal' in iptal_durum.lower()

            tax_inclusive = toplam_tutar + vergiler_toplami

            rows.append({
                'invoice_id': invoice_id,
                'sender': sender_vkn,
                'sender_name': sender_name,
                'issue_date': issue_date,
                'tax_exclusive_amount': toplam_tutar,
                'payable_amount': odenecek_tutar,
                'tax_inclusive_amount': tax_inclusive,
                'currency_code': para_birimi if para_birimi else 'TRY',
                'is_iptal': is_iptal,
            })
        return rows

    def _build_vals(self, row_data, company):
        """Build guven.fatura create/write values from parsed row data."""
        is_try = (row_data['currency_code'] or 'TRY').upper() == 'TRY'
        is_iptal = row_data['is_iptal']

        vals = {
            'invoice_id': row_data['invoice_id'],
            'sender': row_data['sender'],
            'sender_name': row_data['sender_name'],
            'issue_date': row_data['issue_date'],
            'currency_code': row_data['currency_code'],
            'tax_exclusive_amount': row_data['tax_exclusive_amount'],
            'tax_inclusive_amount': row_data['tax_inclusive_amount'],
            'payable_amount': row_data['payable_amount'],
            'exchange_rate': 1.0,
            'kaynak': 'e-arsiv-gibexcel',
            'direction': 'IN',
            'profile_id': 'EARSIVFATURA',
            'details_received': True,
            'is_cancellation': False,
            'receiver': company.vat or '',
            'receiver_name': company.name or '',
            'company_id': company.id,
            'harici_iptal': is_iptal,
            'status_code': '150' if is_iptal else '130',
        }
        if is_try:
            vals['tax_exclusive_amount_try'] = row_data['tax_exclusive_amount']
            vals['tax_inclusive_amount_try'] = row_data['tax_inclusive_amount']
            vals['payable_amount_try'] = row_data['payable_amount']

        return vals

    # ----------------------------------------------------------------
    # Actions
    # ----------------------------------------------------------------

    def action_preview(self):
        """Validate files and compute preview counts (draft → preview)."""
        self.ensure_one()
        if not self.file_ids:
            raise UserError("Lütfen en az bir Excel dosyası ekleyin.")

        Fatura = self.env['guven.fatura']
        company = self.company_id
        total = new = update = skip = 0

        for frec in self.file_ids:
            wb = self._read_workbook(frec)
            ws = wb.active
            self._validate_headers(ws, frec.file_name or 'dosya')
            rows = self._parse_rows(ws)
            wb.close()

            file_new = file_update = file_skip = 0
            for row_data in rows:
                existing = Fatura.search([
                    ('invoice_id', '=', row_data['invoice_id']),
                    ('sender', '=', row_data['sender']),
                    ('kaynak', '=', 'e-arsiv-gibexcel'),
                    ('company_id', '=', company.id),
                ], limit=1)
                if existing:
                    if existing.is_locked:
                        file_skip += 1
                    else:
                        file_update += 1
                else:
                    file_new += 1

            frec.write({
                'row_count': len(rows),
                'status': f"{file_new} yeni, {file_update} güncelleme, {file_skip} atlama",
            })
            total += len(rows)
            new += file_new
            update += file_update
            skip += file_skip

        self.write({
            'state': 'preview',
            'preview_total': total,
            'preview_new': new,
            'preview_update': update,
            'preview_skip': skip,
        })
        return {
            'type': 'ir.actions.act_window',
            'res_model': self._name,
            'res_id': self.id,
            'view_mode': 'form',
            'target': 'new',
        }

    def action_import(self):
        """Execute import (preview → done)."""
        self.ensure_one()
        Fatura = self.env['guven.fatura']
        company = self.company_id
        created = updated = skipped = 0
        errors = []

        for frec in self.file_ids:
            wb = self._read_workbook(frec)
            ws = wb.active
            rows = self._parse_rows(ws)
            wb.close()

            for row_data in rows:
                try:
                    existing = Fatura.search([
                        ('invoice_id', '=', row_data['invoice_id']),
                        ('sender', '=', row_data['sender']),
                        ('kaynak', '=', 'e-arsiv-gibexcel'),
                        ('company_id', '=', company.id),
                    ], limit=1)

                    vals = self._build_vals(row_data, company)

                    if existing:
                        if existing.is_locked:
                            skipped += 1
                            continue
                        # Don't overwrite uuid on update
                        existing.write(vals)
                        updated += 1
                    else:
                        uuid = f"{row_data['invoice_id']}_{row_data['sender']}_GIBEXCEL"
                        vals['uuid'] = uuid
                        Fatura.create(vals)
                        created += 1
                except Exception as e:
                    _logger.exception(
                        "E-Arşiv GİB import error for invoice %s",
                        row_data.get('invoice_id', '?'),
                    )
                    errors.append(f"{row_data.get('invoice_id', '?')}: {e}")

        # Build summary HTML
        summary_parts = [
            '<div class="alert alert-success">',
            f'<strong>Import tamamlandı!</strong><br/>',
            f'Yeni: <strong>{created}</strong> | ',
            f'Güncellenen: <strong>{updated}</strong> | ',
            f'Atlanan (kilitli): <strong>{skipped}</strong>',
            '</div>',
        ]
        if errors:
            summary_parts.append('<div class="alert alert-danger">')
            summary_parts.append(f'<strong>{len(errors)} hata:</strong><ul>')
            for err in errors[:20]:
                summary_parts.append(f'<li>{err}</li>')
            if len(errors) > 20:
                summary_parts.append(f'<li>...ve {len(errors) - 20} hata daha</li>')
            summary_parts.append('</ul></div>')

        # Optional Logo sync
        logo_msg = ''
        if self.auto_logo_sync and created + updated > 0:
            try:
                logo_wizard = self.env['guven.logo.sync.wizard'].create({
                    'company_ids': [(6, 0, [company.id])],
                })
                logo_wizard.action_sync()
                match_total = logo_wizard.match_total or 0
                match_single = logo_wizard.match_single or 0
                logo_msg = (
                    f'<div class="alert alert-info">'
                    f'<strong>Logo Eşleştirme:</strong> '
                    f'{match_total} fatura tarandı, {match_single} eşleşme bulundu.'
                    f'</div>'
                )
            except Exception as e:
                _logger.exception("Logo sync after GIB Excel import failed")
                logo_msg = (
                    f'<div class="alert alert-warning">'
                    f'<strong>Logo eşleştirme hatası:</strong> {e}'
                    f'</div>'
                )

        self.write({
            'state': 'done',
            'import_summary': ''.join(summary_parts) + logo_msg,
        })
        return {
            'type': 'ir.actions.act_window',
            'res_model': self._name,
            'res_id': self.id,
            'view_mode': 'form',
            'target': 'new',
        }

    def action_back(self):
        """Go back to draft state (preview → draft)."""
        self.ensure_one()
        self.write({
            'state': 'draft',
            'preview_total': 0,
            'preview_new': 0,
            'preview_update': 0,
            'preview_skip': 0,
        })
        return {
            'type': 'ir.actions.act_window',
            'res_model': self._name,
            'res_id': self.id,
            'view_mode': 'form',
            'target': 'new',
        }

    def action_close(self):
        return {
            'type': 'ir.actions.client',
            'tag': 'reload',
        }
